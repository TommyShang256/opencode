"""Indexer - 构建DIS知识库索引

从数据字典(.xlsm/.xlsx)解析话单和字段定义，生成JSON索引文件。

支持两种格式：
1. DIS格式 (.xlsm)：话单定义为Sheet，每个Sheet对应一个话单
2. Smart Optimization格式 (.xlsx)：事件目录为Sheet2，各协议事件为独立Sheet

索引结构：
{
  "metadata": { ... },
  "records": { 话单名称: { fields: [...], total_fields: n } },
  "fields": { "话单名称|字段名": { ... } },
  "keywords": { 关键词: ["话单名称|字段名", ...] }
}
"""
import os
import sys
import json
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Set, Optional
from datetime import datetime
import jieba

# 设置输出编码为UTF-8以支持中文显示
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


def get_shared_strings(xlsm_path: str) -> List[str]:
    """从xlsm/xlsx文件读取共享字符串表"""
    strings = []
    try:
        with zipfile.ZipFile(xlsm_path, 'r') as z:
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                for si in root.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    text = ''
                    for t in si.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                        text += (t.text or '')
                    strings.append(text)
    except Exception as e:
        print(f"警告: 无法读取共享字符串 - {e}")
    return strings


def parse_xlsx_direct(xlsx_path: str) -> Dict:
    """直接解析xlsx文件（使用openpyxl，支持中文Sheet名）

    解析逻辑：
    - 每个Sheet可能包含多张话单
    - 通过 "Table Name" 行来识别不同话单的边界
    - 格式：Table Name | 话单名称 | (空) - 表示新话单开始
    - 下一行：字段类型 | 数据库字段名 | 字段中文名 | 数据类型 | 单位 | 取值范围 | 字段说明
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        result = {}
        # 遍历所有sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 当前话单状态
            current_table_name = None
            current_fields = []
            in_header = False

            for row in ws.iter_rows(values_only=True):
                row = list(row)

                # 跳过完全空的行
                if not row or not any(row for x in row if x):
                    continue

                first_cell = str(row[0]) if row[0] else ''

                # 检查是否是新话单的Table Name行
                if 'Table Name' in first_cell:
                    # 保存上一个话单（如果存在）
                    if current_table_name and current_fields:
                        result[current_table_name] = {
                            '话单名称': current_table_name,
                            '原始Sheet': sheet_name,
                            '字段数': len(current_fields),
                            '字段列表': current_fields
                        }

                    # 开始新话单
                    current_table_name = str(row[1]) if row[1] else ''
                    current_fields = []
                    continue

                # 如果没有正在解析的话单，跳过
                if not current_table_name:
                    continue

                # 检查是否是字段定义的表头行（包含"字段类型"文字的行）
                if '字段类型' in first_cell:
                    # 这行是表头，跳过
                    continue

                # 解析字段定义
                # 所有以 dimension/fact/counter 开头的行都是字段
                if first_cell and first_cell.lower() in ['dimension', 'fact', 'counter']:
                    if len(row) >= 5 and row[1]:  # 需要有数据库字段名
                        field_name = str(row[1]) if row[1] else ''
                        # 跳过表头中的字段名
                        if field_name in ['数据库字段名称', 'Field Name', 'FieldName', '数据库字段名']:
                            continue

                        field = {
                            '字段类型': row[0] if len(row) > 0 else '',
                            '数据库字段名': field_name,
                            '字段中文名': str(row[2]) if len(row) > 2 and row[2] else '',
                            '字段类型_DB': str(row[3]) if len(row) > 3 and row[3] else '',
                            '单位': str(row[4]) if len(row) > 4 and row[4] else '',
                            '取值范围': str(row[5]) if len(row) > 5 and row[5] else '',
                            '字段说明': str(row[6]) if len(row) > 6 and row[6] else ''
                        }
                        current_fields.append(field)

            # 保存最后一个话单
            if current_table_name and current_fields:
                result[current_table_name] = {
                    '话单名称': current_table_name,
                    '原始Sheet': sheet_name,
                    '字段数': len(current_fields),
                    '字段列表': current_fields
                }

        return result
    except ImportError:
        print("警告: openpyxl未安装，无法解析xlsx格式")
        return {}
    except Exception as e:
        print(f"警告: 解析xlsx '{xlsx_path}' 失败 - {e}")
        import traceback
        traceback.print_exc()
        return {}


def parse_iom_xlsx_direct(xlsx_path: str) -> Dict:
    """解析IOM格式的xlsx文件（使用openpyxl）

    IOM格式特征：
    - 话单分隔符：'表名'（不是 'Table Name'）
    - 表头：['字段类型', '数据库字段名称', '字段名称', '数据类型', '单位']
    - 缺少 '取值范围' 和 '英文字段名' 列

    解析逻辑：
    - 通过 "表名" 行来识别不同话单的边界
    - 格式：表名 | 话单名称 | (空) - 表示新话单开始
    - 下一行：字段类型 | 数据库字段名称 | 字段名称 | 数据类型 | 单位
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        result = {}
        # 遍历所有sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 当前话单状态
            current_table_name = None
            current_fields = []
            in_header = False

            for row in ws.iter_rows(values_only=True):
                row = list(row)

                # 跳过完全空的行
                if not row or not any(row for x in row if x):
                    continue

                first_cell = str(row[0]) if row[0] else ''

                # 检查是否是新话单的"表名"行（IOM格式特征）
                if '表名' in first_cell:
                    # 保存上一个话单（如果存在）
                    if current_table_name and current_fields:
                        result[current_table_name] = {
                            '话单名称': current_table_name,
                            '原始Sheet': sheet_name,
                            '字段数': len(current_fields),
                            '字段列表': current_fields,
                            'is_iom_format': True
                        }

                    # 开始新话单
                    current_table_name = str(row[1]) if row[1] else ''
                    current_fields = []
                    continue

                # 如果没有正在解析的话单，跳过
                if not current_table_name:
                    continue

                # 检查是否是字段定义的表头行（IOM格式表头）
                if '字段类型' in first_cell and '数据库字段名称' in str(row[1] if len(row) > 1 else ''):
                    # 这行是表头，跳过
                    continue

                # 解析字段定义
                # 所有以 dimension/fact/counter 开头的行都是字段
                if first_cell and first_cell.lower() in ['dimension', 'fact', 'counter']:
                    if len(row) >= 4 and row[1]:  # 需要有数据库字段名
                        field_name = str(row[1]) if row[1] else ''
                        # 跳过表头中的字段名
                        if field_name in ['数据库字段名称', 'Field Name', 'FieldName', '数据库字段名']:
                            continue

                        field = {
                            '字段类型': row[0] if len(row) > 0 else '',
                            '数据库字段名': field_name,
                            '字段中文名': str(row[2]) if len(row) > 2 and row[2] else '',
                            '字段类型_DB': str(row[3]) if len(row) > 3 and row[3] else '',
                            '单位': str(row[4]) if len(row) > 4 and row[4] else '',
                            '取值范围': '',  # IOM格式没有此列
                            '字段说明': str(row[2]) if len(row) > 2 and row[2] else ''  # 字段说明用中文名填充
                        }
                        current_fields.append(field)

            # 保存最后一个话单
            if current_table_name and current_fields:
                result[current_table_name] = {
                    '话单名称': current_table_name,
                    '原始Sheet': sheet_name,
                    '字段数': len(current_fields),
                    '字段列表': current_fields,
                    'is_iom_format': True
                }

        return result
    except ImportError:
        print("警告: openpyxl未安装，无法解析xlsx格式")
        return {}
    except Exception as e:
        print(f"警告: 解析IOM xlsx '{xlsx_path}' 失败 - {e}")
        import traceback
        traceback.print_exc()
        return {}


def parse_xlsm_sheet(xlsm_path: str, sheet_name: str, shared_strings: List[str]) -> Optional[Dict]:
    """解析xlsm中的指定sheet（DIS格式）"""
    try:
        with zipfile.ZipFile(xlsm_path, 'r') as z:
            # 找到sheet对应的文件
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                # 找到sheet对应的rId
                sheet_rid = None
                for sheet in root.findall('.//ns:sheet', ns):
                    if sheet.get('name') == sheet_name:
                        sheet_rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                        break

                if not sheet_rid:
                    return None

                # 从relationships找到实际文件
                with z.open('xl/_rels/workbook.xml.rels') as f:
                    rels_tree = ET.parse(f)
                    rels_root = rels_tree.getroot()
                    sheet_file = None
                    for rel in rels_root:
                        if rel.get('Id') == sheet_rid:
                            sheet_file = rel.get('Target')
                            break

                    if not sheet_file:
                        return None

                    sheet_file = f'xl/{sheet_file}' if not sheet_file.startswith('xl/') else sheet_file

                    # 解析sheet内容
                    with z.open(sheet_file) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                        fields = []
                        for row_idx, row in enumerate(root.findall('.//ns:row', ns)):
                            if row_idx == 0:
                                continue  # 跳过表头

                            cells = []
                            for cell in row.findall('ns:c', ns):
                                value_elem = cell.find('ns:v', ns)
                                value = ''
                                if value_elem is not None and value_elem.text:
                                    cell_type = cell.get('t', '')
                                    if cell_type == 's':
                                        idx = int(value_elem.text)
                                        value = shared_strings[idx] if idx < len(shared_strings) else ''
                                    else:
                                        value = value_elem.text
                                cells.append(value)

                            # 跳过表头行（第一行）
                            if row_idx == 1:
                                continue

                            if len(cells) >= 5 and cells[1]:  # 需要有字段名
                                # 跳过可能是表头的行
                                field_name = cells[1] if len(cells) > 1 else ''
                                if '字段名称' in field_name or 'FieldName' in field_name:
                                    continue

                                field = {
                                    '序号': cells[0] if len(cells) > 0 else '',
                                    '数据库字段名': cells[1] if len(cells) > 1 else '',
                                    '英文字段名': cells[2] if len(cells) > 2 else '',
                                    '字段中文名': cells[3] if len(cells) > 3 else '',
                                    '字段类型': cells[4] if len(cells) > 4 else '',
                                    '字段含义': cells[5] if len(cells) > 5 else '',
                                    '英文描述': cells[6] if len(cells) > 6 else '',
                                    '是否可见': cells[7] if len(cells) > 7 else 'N'
                                }
                                fields.append(field)

                        return {
                            '话单名称': sheet_name,
                            '字段数': len(fields),
                            '字段列表': fields
                        }
    except Exception as e:
        print(f"警告: 解析sheet '{sheet_name}' 失败 - {e}")
        return None


class DisIndexer:
    """DIS知识库索引器"""

    def __init__(self, knowledge_dir: str):
        self.knowledge_dir = Path(knowledge_dir)
        self.fields_dir = self.knowledge_dir / "fields"

        # 索引数据结构
        self.records: Dict[str, Dict] = {}      # 话单名称 -> 话单信息
        self.keywords: Dict[str, Set[str]] = {} # 关键词 -> 字段标识集合

    def build_index(self, copy_from_dir: str = None) -> Dict:
        """构建索引

        Args:
            copy_from_dir: 如果指定，复制该目录下的数据字典文件到knowledge目录
        """
        if not self.knowledge_dir.exists():
            self.knowledge_dir.mkdir(parents=True, exist_ok=True)

        # 创建fields子目录
        self.fields_dir.mkdir(exist_ok=True)

        # 如果指定了复制源目录
        if copy_from_dir:
            copy_dir = Path(copy_from_dir)
            if not copy_dir.exists():
                raise ValueError(f"数据字典目录不存在: {copy_from_dir}")

            print(f"从 {copy_from_dir} 复制数据字典文件...")
            # 复制所有支持的数据字典文件（.xlsm 和 .xlsx）
            for ext in ['*.xlsm', '*.xlsx']:
                for dict_file in copy_dir.glob(ext):
                    if dict_file.name.startswith('~$'):
                        continue
                    dest_path = self.knowledge_dir / dict_file.name
                    shutil.copy2(dict_file, dest_path)
                    print(f"  复制: {dict_file.name}")

        # 查找所有数据字典文件（.xlsm 和 .xlsx，跳过临时文件）
        dict_files = []
        for ext in ['*.xlsm', '*.xlsx']:
            dict_files.extend([f for f in self.knowledge_dir.glob(ext) if not f.name.startswith('~$')])
        print(f"找到 {len(dict_files)} 个数据字典文件")

        if not dict_files:
            raise ValueError("未找到数据字典文件(.xlsm/.xlsx)，请先放入knowledge目录")

        # 解析每个数据字典文件
        for dict_file in dict_files:
            print(f"\n解析: {dict_file.name}")

            # 根据文件类型选择解析方式
            if dict_file.suffix.lower() == '.xlsx':
                # 检查是否为IOM格式文件（文件名包含IOM）
                if 'IOM' in dict_file.name.upper():
                    # IOM格式 - 使用专门的IOM解析器
                    print(f"  检测到IOM格式文件，使用IOM解析器...")
                    records = parse_iom_xlsx_direct(str(dict_file))
                    for record_name, record_info in records.items():
                        self._add_record(record_info)
                        print(f"  [OK] {record_name}: {record_info['字段数']} fields")
                else:
                    # Smart Optimization格式 - 使用openpyxl直接解析
                    records = parse_xlsx_direct(str(dict_file))
                    for record_name, record_info in records.items():
                        self._add_record(record_info)
                        print(f"  [OK] {record_name}: {record_info['字段数']} fields")
            else:
                # DIS格式 - 使用zipfile解析
                shared_strings = get_shared_strings(str(dict_file))

                with zipfile.ZipFile(str(dict_file), 'r') as z:
                    with z.open('xl/workbook.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        sheets = root.findall('.//ns:sheet', ns)

                        for sheet in sheets:
                            sheet_name = sheet.get('name')
                            # 跳过非话单sheet（如目录、说明等）
                            if any(skip in sheet_name for skip in ['说明', '定义', '目录', '规则', '列表']):
                                if '话单定义' not in sheet_name and '详单定义' not in sheet_name:
                                    continue

                            record_info = parse_xlsm_sheet(str(dict_file), sheet_name, shared_strings)
                            if record_info and record_info['字段列表']:
                                self._add_record(record_info)
                                print(f"  [OK] {sheet_name}: {record_info['字段数']} fields")

        # 保存索引
        result = self._save_index(dict_files)

        print(f"\n索引构建完成:")
        print(f"  - 话单数量: {len(self.records)}")
        print(f"  - 字段目录: {self.fields_dir}")

        return result

    def _add_record(self, record_info: Dict):
        """添加话单到索引"""
        record_name = record_info['话单名称']
        fields = record_info['字段列表']

        # 存储话单信息
        self.records[record_name] = {
            '话单名称': record_name,
            '原始Sheet': record_info.get('原始Sheet', ''),
            '字段数': len(fields),
            '字段列表': [f['数据库字段名'] for f in fields],
            '字段文件': f"fields/{record_name}.json"
        }

        # L1.5话单（如L15_NR_XXX）还需要添加HDFS版本的索引（去掉L15_前缀）
        if record_name.startswith('L15_'):
            hdfs_name = record_name[4:]  # 去掉 "L15_"
            self.records[hdfs_name] = {
                '话单名称': hdfs_name,
                '原始话单': record_name,
                '原始Sheet': record_info.get('原始Sheet', ''),
                '字段数': len(fields),
                '字段列表': [f['数据库字段名'] for f in fields],
                '字段文件': f"fields/{hdfs_name}.json",
                'is_hdfs_alias': True
            }
            # 为HDFS版本话单也提取关键词
            self._extract_keywords_for_record(hdfs_name, fields)

        # 提取关键词（只为话单名称提取，因为字段信息在单独文件中）
        self._extract_keywords_for_record(record_name, fields)

    def _extract_keywords_for_record(self, record_name: str, fields: List[Dict]):
        """为话单提取关键词（用于话单级别搜索）"""
        keywords = set()

        # 话单名称
        keywords.add(record_name.lower())
        for word in re.findall(r'[A-Za-z]+', record_name):
            keywords.add(word.lower())

        # 更新倒排索引
        for kw in keywords:
            if kw not in self.keywords:
                self.keywords[kw] = set()
            self.keywords[kw].add(record_name)

    def _generate_aliases(self) -> Dict[str, str]:
        """生成话单别名映射

        将话单生成对应的别名（去掉前缀）
        例如：L15_NR_UE_MR → NR_UE_MR
             F_SERVEXP_ASS_CELL_GRID_H → SERVEXP_ASS_CELL_GRID_H

        Returns:
            别名字典 { 别名: 标准话单名 }
        """
        aliases = {}

        for record_name in self.records.keys():
            # 处理 L15_ 前缀的话单
            if record_name.startswith('L15_'):
                alias = record_name[4:]  # 去掉前4个字符 "L15_"
                aliases[alias] = record_name

            # 处理 F_ 前缀的话单（L2话单）
            elif record_name.startswith('F_'):
                alias = record_name[2:]  # 去掉前2个字符 "F_"
                aliases[alias] = record_name

            # 处理 f_ 前缀的话单（小写）
            elif record_name.startswith('f_'):
                alias = record_name[2:]  # 去掉前2个字符 "f_"
                aliases[alias] = record_name
                # 同时生成大写版本
                alias_upper = alias.upper()
                if alias_upper != alias:
                    aliases[alias_upper] = record_name

        return aliases

    def _save_index(self, dict_files: List[Path]) -> Dict:
        """保存索引文件"""
        # 转换set为list用于JSON序列化
        keywords_list = {k: list(v) for k, v in self.keywords.items()}

        # 生成话单别名映射（L15_NR_UE_MR → NR_UE_MR）
        aliases = self._generate_aliases()

        # 保存主索引（只包含话单列表和字段名）
        manifest = {
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "total_records": len(self.records),
                "total_with_aliases": len(self.records) + len(aliases),
                "knowledge_path": str(self.knowledge_dir),
                "split_index": True  # 标记为拆分索引
            },
            "records": self.records,
            "aliases": aliases,  # 别名映射
            "keywords": keywords_list
        }

        manifest_path = self.knowledge_dir / "index_manifest.json"
        with open(manifest_path, 'w', encoding='utf-8') as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)

        # 单独保存每个话单的字段详情到fields目录
        # L1.5话单（以L15_开头）会生成两个版本：Carbon表(L15_XXX.json)和HDFS文件(XXX.json)
        print(f"\n保存字段详情文件...")
        for dict_file in dict_files:
            if dict_file.suffix.lower() == '.xlsx':
                # 检查是否为IOM格式文件
                if 'IOM' in dict_file.name.upper():
                    # IOM格式
                    records = parse_iom_xlsx_direct(str(dict_file))
                    for record_name, record_info in records.items():
                        self._save_fields_file(record_name, record_info['字段列表'], is_xlsx=True, is_iom=True)
                else:
                    # Smart Optimization格式
                    records = parse_xlsx_direct(str(dict_file))
                    for record_name, record_info in records.items():
                        self._save_fields_file(record_name, record_info['字段列表'], is_xlsx=True)
            else:
                # DIS格式
                shared_strings = get_shared_strings(str(dict_file))
                with zipfile.ZipFile(str(dict_file), 'r') as z:
                    with z.open('xl/workbook.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        sheets = root.findall('.//ns:sheet', ns)

                        for sheet in sheets:
                            sheet_name = sheet.get('name')
                            if any(skip in sheet_name for skip in ['说明', '定义', '目录', '规则', '列表']):
                                if '话单定义' not in sheet_name and '详单定义' not in sheet_name:
                                    continue

                            record_info = parse_xlsm_sheet(str(dict_file), sheet_name, shared_strings)
                            if record_info and record_info['字段列表']:
                                self._save_fields_file(sheet_name, record_info['字段列表'], is_xlsx=False)

        return {
            "manifest_path": str(manifest_path),
            "fields_dir": str(self.fields_dir),
            "total_records": len(self.records)
        }

    def _save_fields_file(self, record_name: str, fields: List[Dict], is_xlsx: bool = False, is_iom: bool = False):
        """保存字段详情到单独文件

        L1.5话单（如L15_NR_XXX）需要生成两个版本：
        - Carbon表版本：L15_NR_XXX.json
        - HDFS文件版本：NR_XXX.json（去掉L15_前缀）

        Args:
            record_name: 话单名称
            fields: 字段列表
            is_xlsx: 是否为Smart Optimization格式
            is_iom: 是否为IOM格式
        """
        fields_data = self._build_fields_data(record_name, fields, is_xlsx, is_iom)

        # 判断是否需要生成HDFS版本（话单名称以L15_开头）
        if record_name.startswith('L15_'):
            # 生成Carbon表版本（原名称）
            field_file = self.fields_dir / f"{record_name}.json"
            with open(field_file, 'w', encoding='utf-8') as f:
                json.dump(fields_data, f, ensure_ascii=False, indent=2)

            # 生成HDFS文件版本（去掉L15_前缀）
            hdfs_name = record_name[4:]  # 去掉 "L15_"
            field_file_hdfs = self.fields_dir / f"{hdfs_name}.json"
            with open(field_file_hdfs, 'w', encoding='utf-8') as f:
                json.dump(fields_data, f, ensure_ascii=False, indent=2)
        else:
            # 其他话单正常生成一个版本
            field_file = self.fields_dir / f"{record_name}.json"
            with open(field_file, 'w', encoding='utf-8') as f:
                json.dump(fields_data, f, ensure_ascii=False, indent=2)

    def _build_fields_data(self, record_name: str, fields: List[Dict], is_xlsx: bool, is_iom: bool = False) -> List[Dict]:
        """构建字段数据列表

        Args:
            record_name: 话单名称
            fields: 字段列表
            is_xlsx: 是否为Smart Optimization格式
            is_iom: 是否为IOM格式
        """
        fields_data = []
        for field in fields:
            if is_iom:
                # IOM格式字段（缺少英文字段名和取值范围）
                fields_data.append({
                    '话单名称': record_name,
                    '字段类型': field.get('字段类型', ''),
                    '数据库字段名': field.get('数据库字段名', ''),
                    '字段中文名': field.get('字段中文名', ''),
                    '字段类型_DB': field.get('字段类型_DB', ''),
                    '单位': field.get('单位', ''),
                    '取值范围': field.get('取值范围', ''),
                    '字段说明': field.get('字段说明', ''),
                    'is_iom_format': True
                })
            elif is_xlsx:
                # Smart Optimization格式字段
                fields_data.append({
                    '话单名称': record_name,
                    '字段类型': field.get('字段类型', ''),
                    '数据库字段名': field.get('数据库字段名', ''),
                    '字段中文名': field.get('字段中文名', ''),
                    '字段类型_DB': field.get('字段类型_DB', ''),
                    '单位': field.get('单位', ''),
                    '取值范围': field.get('取值范围', ''),
                    '字段说明': field.get('字段说明', '')
                })
            else:
                # DIS格式字段
                fields_data.append({
                    '话单名称': record_name,
                    '数据库字段名': field.get('数据库字段名', ''),
                    '英文字段名': field.get('英文字段名', ''),
                    '字段中文名': field.get('字段中文名', ''),
                    '字段类型': field.get('字段类型', ''),
                    '字段含义': field.get('字段含义', ''),
                    '英文描述': field.get('英文描述', ''),
                    '是否可见': field.get('是否可见', 'N')
                })
        return fields_data


def build_index(knowledge_dir: str, copy_from_dir: str = None) -> Dict:
    """构建索引的入口函数

    Args:
        knowledge_dir: 知识库目录（存放xlsm和输出json）
        copy_from_dir: 可选，从该目录复制xlsm文件到knowledge_dir

    Returns:
        索引构建结果
    """
    indexer = DisIndexer(knowledge_dir)
    return indexer.build_index(copy_from_dir)


if __name__ == "__main__":
    # 获取技能目录（向上两级）
    script_dir = Path(__file__).parent
    skill_dir = script_dir.parent
    knowledge_dir = skill_dir / "knowledge"

    # 解析命令行参数
    copy_from_dir = None
    if len(sys.argv) > 1:
        copy_from_dir = sys.argv[1]

    print("=" * 60)
    print("DIS Indexer - 构建知识库索引")
    print("=" * 60)
    print(f"技能目录: {skill_dir}")
    print(f"知识库目录: {knowledge_dir}")
    if copy_from_dir:
        print(f"数据字典来源: {copy_from_dir}")
    print("=" * 60)

    result = build_index(str(knowledge_dir), copy_from_dir=copy_from_dir)
    print(f"\n✅ 索引构建成功!")
    print(f"   - 话单数量: {result['total_records']}")
    print(f"   - 主索引: {result['manifest_path']}")
    print(f"   - 字段目录: {result['fields_dir']}")