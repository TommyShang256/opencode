"""SubscriptionParser - 解析DSI订阅Excel文件

解析DSI推送的订阅Excel，获取：
1. Subscription List - 订阅配置信息
2. Data Record Field - 订阅字段列表
"""
import os
import re
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl.cell.cell import MergedCell
import openpyxl


def get_cell_value(cell) -> str:
    """获取单元格值

    支持多种数据类型：
    1. cell对象：cell.value
    2. 字符串/数值：直接转换
    3. None：返回空字符串
    """
    # None值
    if cell is None:
        return ""

    # 如果是cell对象
    if hasattr(cell, 'value'):
        if isinstance(cell, MergedCell):
            return ""
        value = cell.value
        if value is None:
            return ""
        value = str(value).replace("\n", " ").replace("\r", " ")
        value = " ".join(value.split())
        return value.strip()

    # 其他类型（字符串、数值等）
    try:
        value = str(cell).replace("\n", " ").replace("\r", " ")
        value = " ".join(value.split())
        return value.strip()
    except:
        return ""


class SubscriptionParser:
    """订阅Excel解析器"""

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")

        self.wb = openpyxl.load_workbook(excel_path, data_only=True)

        # 解析结果
        self.subscriptions: List[Dict] = []      # 订阅列表
        self.fields: List[Dict] = {}             # 话单名称 -> 字段列表

    def parse(self) -> Dict:
        """解析整个Excel文件"""
        # 1. 解析 Subscription List
        self._parse_subscription_list()

        # 2. 解析 Data Record Field
        self._parse_data_record_field()

        return {
            'subscriptions': self.subscriptions,
            'fields': self.fields
        }

    def get_subscription(self, name: str) -> Optional[Dict]:
        """获取指定订阅详情"""
        for sub in self.subscriptions:
            if sub.get('订阅名称') == name:
                return sub
        return None

    def get_fields_by_subscription(self, subscription_name: str) -> List[Dict]:
        """获取指定订阅的所有字段"""
        return self.fields.get(subscription_name, [])

    def get_subscription_names(self) -> List[str]:
        """获取所有订阅名称"""
        return [sub.get('订阅名称', '') for sub in self.subscriptions]

    def _parse_subscription_list(self):
        """解析Subscription List sheet"""
        sheet_name = 'Subscription List'

        if sheet_name not in self.wb.sheetnames:
            # 尝试模糊匹配
            for name in self.wb.sheetnames:
                if 'subscription' in name.lower() or '订阅' in name:
                    sheet_name = name
                    break
            else:
                return

        ws = self.wb[sheet_name]

        # 解析表头
        headers = []
        data_rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # 表头行
                headers = [get_cell_value(cell) for cell in row]
                headers = [h for h in headers if h]  # 去除空列
            else:
                # 数据行
                values = [get_cell_value(cell) for cell in row]
                if any(values):  # 非空行
                    # 确保列数一致
                    while len(values) < len(headers):
                        values.append('')
                    data_rows.append(values)

        # 转换为字典列表
        for row in data_rows:
            if len(row) >= 2 and row[1]:  # 订阅名称不为空
                sub_data = dict(zip(headers, row))
                self.subscriptions.append(sub_data)

    def _parse_data_record_field(self):
        """解析Data Record Field sheet"""
        sheet_name = 'Data Record Field'

        if sheet_name not in self.wb.sheetnames:
            # 尝试模糊匹配
            for name in self.wb.sheetnames:
                if 'field' in name.lower() or '字段' in name:
                    sheet_name = name
                    break
            else:
                return

        ws = self.wb[sheet_name]

        # 解析表头
        headers = []
        data_rows = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # 表头行
                headers = [get_cell_value(cell) for cell in row]
                headers = [h for h in headers if h]
            else:
                # 数据行
                values = [get_cell_value(cell) for cell in row]
                if any(values):
                    while len(values) < len(headers):
                        values.append('')
                    data_rows.append(values)

        # 按订阅分组
        for row in data_rows:
            if len(row) >= 3:  # 至少要有订阅名称、话单名称、字段名
                field_data = dict(zip(headers, row))
                subscription_name = field_data.get('订阅名称', '')
                if subscription_name:
                    if subscription_name not in self.fields:
                        self.fields[subscription_name] = []
                    self.fields[subscription_name].append(field_data)


def parse_subscription_excel(excel_path: str) -> Dict:
    """解析订阅Excel的便捷函数"""
    parser = SubscriptionParser(excel_path)
    return parser.parse()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: subscription_parser.py <订阅Excel路径>")
        sys.exit(1)

    excel_path = sys.argv[1]
    result = parse_subscription_excel(excel_path)

    print(f"解析完成:")
    print(f"  订阅数量: {len(result['subscriptions'])}")
    print(f"  订阅列表: {[s.get('订阅名称') for s in result['subscriptions']]}")
    print()

    for sub_name, fields in result['fields'].items():
        print(f"  {sub_name}: {len(fields)} 个字段")